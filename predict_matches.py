import sys
sys.stdout.reconfigure(encoding='utf-8')
import pandas as pd
import joblib
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load model and scaler
model = joblib.load("football_model_optimized.pkl")
scaler = joblib.load("scaler.pkl")

# Load Excel file
df = pd.read_excel("new_matches.xlsx")

# Stake mapping
stake_map = {"Small": 1, "Medium": 2, "Large": 3}

# Score parser
def parse_score(score):
    if isinstance(score, str):
        score = score.replace(" ", "")
        parts = score.split("-")
        if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
            return int(parts[0]), int(parts[1])
    return 0, 0

# Helper for vote tally
def vote_label(h, a):
    if h > a:
        return 1
    elif h < a:
        return -1
    return 0

processed_data = []
predicted_scores = []
skipped_rows = []
teams = []

for idx, row in df.iterrows():
    try:
        p1_h, p1_a = parse_score(row["Prediction 1"])
        p2_h, p2_a = parse_score(row["Prediction 2"])
        p3_h, p3_a = parse_score(row["Prediction 3"])
        p4_h, p4_a = parse_score(row["Prediction 4"])
        stake = stake_map.get(str(row["Stake"]).strip().title(), 2)
        teams.append(row["Teams"])

        # Avg score for display
        avg_home = round((p1_h + p2_h + p3_h + p4_h) / 4)
        avg_away = round((p1_a + p2_a + p3_a + p4_a) / 4)
        predicted_scores.append(f"{avg_home}-{avg_away}")

        # Feature Engineering (must match training: 18 features)
        p3_goal_diff = p3_h - p3_a
        p3_total_goals = p3_h + p3_a
        stake_weighted_diff = p3_goal_diff * stake
        p3_vs_others_diff = p3_goal_diff - ((p1_h - p1_a) + (p2_h - p2_a) + (p4_h - p4_a)) / 3
        avg_goal_diff = ((p1_h - p1_a) + (p2_h - p2_a) + (p3_h - p3_a) + (p4_h - p4_a)) / 4
        pred_variance = pd.Series([p1_h, p2_h, p3_h, p4_h]).var() + pd.Series([p1_a, p2_a, p3_a, p4_a]).var()

        # Vote tallies
        votes = [vote_label(p1_h, p1_a), vote_label(p2_h, p2_a), vote_label(p3_h, p3_a), vote_label(p4_h, p4_a)]
        home_win_votes = votes.count(1)
        draw_votes = votes.count(0)
        away_win_votes = votes.count(-1)

        # Final input data row (18 features)
        input_data = [
            p1_h, p1_a, p2_h, p2_a, p3_h, p3_a, p4_h, p4_a,
            stake, p3_goal_diff, p3_total_goals, stake_weighted_diff, p3_vs_others_diff,
            avg_goal_diff, pred_variance,
            home_win_votes, draw_votes, away_win_votes
        ]
        processed_data.append(input_data)

    except Exception as e:
        skipped_rows.append(idx)
        predicted_scores.append("Error")
        teams.append(row["Teams"])
        print(f"⚠️ Skipping row {idx+2} due to error: {e}")

# Convert to DataFrame to ensure correct shape
feature_columns = [
    "P1_H", "P1_A", "P2_H", "P2_A", "P3_H", "P3_A", "P4_H", "P4_A",
    "Stake_num", "P3_goal_diff", "P3_total_goals", "stake_weighted_diff", "P3_vs_others_diff",
    "avg_goal_diff", "pred_variance", "home_win_votes", "draw_votes", "away_win_votes"
]
processed_df = pd.DataFrame(processed_data, columns=feature_columns)

# Verify feature count
assert processed_df.shape[1] == 18, f"Expected 18 features, got {processed_df.shape[1]}"

# Predict
scaled_inputs = scaler.transform(processed_df)
raw_preds = model.predict(scaled_inputs)

# Map to labels
result_map = {-1: "Loss", 0: "Draw", 1: "Win"}
predicted_results = [result_map[p] for p in raw_preds]

# Create output DataFrame with only Teams, Predicted Result, and Predicted Score
output_df = pd.DataFrame({
    "Teams": [teams[i] for i in range(len(teams)) if i not in skipped_rows],
    "Predicted Result": predicted_results,
    "Predicted Score": [score for i, score in enumerate(predicted_scores) if i not in skipped_rows]
})

# Save to Excel
output_file = "predicted_results.xlsx"
output_df.to_excel(output_file, index=False)

# Color-code results
wb = load_workbook(output_file)
ws = wb.active

# Get column index of 'Predicted Result'
result_col_index = None
for i, cell in enumerate(ws[1]):
    if cell.value == "Predicted Result":
        result_col_index = i + 1
        break

# Fill colors
colors = {"Win": "C6EFCE", "Loss": "FFC7CE", "Draw": "D9D9D9"}
if result_col_index:
    for row in range(2, ws.max_row + 1):
        result = ws.cell(row, result_col_index).value
        if result in colors:
            ws.cell(row, result_col_index).fill = PatternFill(
                start_color=colors[result], end_color=colors[result], fill_type="solid"
            )

wb.save(output_file)

# Summary
print("\n✅ Predictions and exact scores saved to 'predicted_results.xlsx' with color-coded results.")
print(f"📊 Summary:\n{output_df['Predicted Result'].value_counts()}")
if skipped_rows:
    print(f"⚠️ Skipped rows (Excel row numbers): {[r+2 for r in skipped_rows]}")
